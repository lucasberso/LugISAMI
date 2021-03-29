import openpyxl
from path import Path
testcases_dir = Path(__file__).dirname()/"testcases"
if __name__ == '__main__':

    filename = 'Lug_manual.xlsm'
    book= openpyxl.load_workbook(testcases_dir+ '/'+ filename)
    sheet = book["Materials"]
    initial_row = 5
    final_row = sheet.max_row
    #print(final_row)
    initial_column = 2
    final_column = sheet.max_column

#lupear
    dict_global = {}
    for i in range(initial_row,final_row+1):

        matname = sheet.cell(i, initial_column).value
        print(matname)

        dict = {}

        for j in range(initial_column, final_column+1):
            keyname = sheet.cell(4,j).value
            value = sheet.cell(i,j).value
            dict.update({keyname:value})

        dict_global.update({matname:dict})

    print(dict_global)

    f = open("myfile.txt", "x")
    f.writelines("########################\n")
    f.writelines("# ISAMI VERSION: 8.0.0 #\n")
    f.writelines("# ANALYSIS: LUG        #\n")
    f.writelines("# Mode: SAA            #\n")
    f.writelines("# Written by: ALTRAN   #\n")
    f.writelines("# Date: 25/01/14       #\n")
    f.writelines("########################\n")

    for keys in dict_global.keys():
        campo1 = keys
        campo2 = dict_global[keys]["ISAMI Name"]
        campo3 = dict_global[keys]["CODE"]
        campo4 = dict_global[keys]["User/Referenced"]
        f.writelines("MS.LoadMaterial("+"'"+ campo1 +"'" +"," + campo2 +"," + campo3 +"," + campo4 +")\n")
