import openpyxl
from path import Path
testcases_dir = Path(__file__).dirname()/"testcases"
if __name__ == '__main__':

    filename = 'Lug_manual.xlsm'
    book= openpyxl.load_workbook(testcases_dir+ '/'+ filename, data_only = True)
    sheet = book["Analysis"]
    initial_row = 4
    final_row = sheet.max_row

    initial_column = 2
    final_column = sheet.max_column - 1

    dict_global = {}
    for i in range(initial_row, final_row + 1):

        Location = sheet.cell(i, 1).value
        print(Location)

        dict = {}

        for j in range(initial_column, final_column + 1):
            keyname = sheet.cell(3, j).value
            value = sheet.cell(i, j).value
            dict.update({keyname: value})

        dict_global.update({Location: dict})

    f = open("myfile.txt", "x")

    cont = 1
    for keys in dict_global.keys():
        campo1 = dict_global[keys]["Standar Pin"]
        campo2 = dict_global[keys]["/Diameter"]
        campo3 = dict_global[keys]["/Material"]
        campo4 = dict_global[keys]["Standar Bush"]
        campo5 = dict_global[keys]["/BushExternalDiameter"]
        campo6 = dict_global[keys]["/BushMaterial"]
        campo7 = dict_global[keys]["/LugResultantForce"]
        campo8 = dict_global[keys]["/LugResultantAngle"]
        campo9 = dict_global[keys]["/LoadingRatio"]
        campo10 = dict_global[keys]["/Width"]
        campo11 = dict_global[keys]["/Length"]
        campo12 = dict_global[keys]["/Thick"]
        campo13 = dict_global[keys]["/ThickUnstudied"]
        campo14 = dict_global[keys]["/PinOffsetX"]
        campo15 = dict_global[keys]["/PinOffsetY"]
        campo16 = dict_global[keys]["/SuperiorAngle"]
        campo17 = dict_global[keys]["/InferiorAngle"]
        campo18 = dict_global[keys]["/ShearType"]
        campo19 = dict_global[keys]["/StructureMaterial"]
        campo20 = dict_global[keys]["/Orientation_init"]



        f.writelines("MS.CreateObject('DPines" + str(cont) + "','AirbusEO_DPin',[\n"
            "['/CsmMbr_Name','S:DPin" + str(cont) + "'],\n"
            "['/CsmMbr_Uptodate','B:TRUE'],\n"
            "['/Diameter','CaesamQty_LENGTH:" + str(campo2) + ";mm'],\n"
            "['/Material','AirbusEO_TMaterial:" + campo3 + "'],\n"
            "])\n\n")

        f.writelines("MS.CreateObject('DBushes" + str(cont) + "','AirbusEO_DBush',[\n"
            "['/CsmMbr_Name','S:D_Bush" + str(cont) + "'],\n"
            "['/BushInternalDiameter','CaesamQty_LENGTH:" + str(campo2) + ";mm'],\n"
            "['/BushExternalDiameter','CaesamQty_LENGTH:" + str(campo5) + ";mm'],\n"
            "['/BushMaterial','AirbusEO_TMaterial:" + campo6 + "'],\n"
            "])\n\n")

        f.writelines("MS.CreateObject('Geom" + str(cont) + "','AirbusEO_DLugGeometry',[\n"
            "['/CsmMbr_Name','S:Geom'],\n"
            "['/LugType','Enum_LugType:SIMPLE'],\n"
            "['/Width','CaesamQty_LENGTH:" + str(campo10) + ";mm'],\n"
            "['/Length','CaesamQty_LENGTH:" + str(campo11) + ";mm'],\n"
            "['/Thick','CaesamQty_LENGTH:" + str(campo12) + ";mm'],\n"
            "['/ThickUnstudied','CaesamQty_LENGTH:" + str(campo13) + ";mm'],\n"
            "['/LugPin','AirbusEO_DPin:DPines" + str(cont) +  "'],\n"
            "['/LugBush','AirbusEO_DBush:DBushes" + str(cont) + "'],\n"
            "['/PinOffsetX','CaesamQty_LENGTH:" + str(campo14) + ";mm'],\n"
            "['/PinOffsetY','CaesamQty_LENGTH:" + str(campo15) + ";mm'],\n"
            "['/SuperiorAngle','CaesamQty_PLANE_ANGLE:" + str(campo16) + ";deg'],\n"
            "['/InferiorAngle','CaesamQty_PLANE_ANGLE:" + str(campo17) + ";deg'],\n"
            "['/NotchedLength','CaesamQty_LENGTH:20;mm'],\n"
            "['/ShearType','Enum_ToggleShearType:"+ campo18 +"'],\n"
            "['/StructureMaterial','AirbusEO_TMaterial:" + campo19 + "'],\n"
            "])\n\n")

        f.writelines("MS.CreateObject('Loading" + str(cont) +"', 'AirbusEO_DLugLoading,[\n"
        "['/CsmMbr_Name','S:Loading'],\n"
        "['/LoadingType','Enum_LoadingType:NO_COEFFICIENT'],\n"
        "['/LugForceLoadingType','Enum_TogglePHForceLoadingType:RESULTANT AND ANGLE'],\n"
        "['/LugCoefficientForceX','D:20'],\n"
        "['/LugCoefficientForceY','D:30'],\n"
        "['/LugForceX','CaesamQty_FORCE:10000;N'],\n"
        "['/LugForceY','CaesamQty_FORCE:10000;N'],\n"
        "['/LugCoefficientResultantForce','D:100'],\n"
        "['/LugResultantForce','CaesamQty_FORCE:" + str(campo7) + ";N'],\n"
        "['/LugResultantAngle','CaesamQty_PLANE_ANGLE:" + str(campo8) + ";deg'],\n"
        "['/NbParameters','I:32'],\n"
        "['/LoadingRatio','D:" + str(campo9) + "'],\n"
        "['/NbClasses','I:1'],\n"
        "['/Compression','Enum_ToggleCompression:Smin'],\n"
        "['/UserSmin','CaesamQty_PRESSURE:0;MPa'],\n"
        "['/PropagationComputation','CaesamEnum_YesNo:No'],\n"
        "['/MLPDesign','CaesamEnum_YesNo:No'],\n"
        "['/LoadRedistributionFactor','D:1']\n"
        "]))\n\n")


        f.writelines("MS.CreateObject('Law" + str(cont) + "','AirbusEO_DFatigueLawGeoDependent',[\n"
        "['/CsmMbr_Name','S:Law'],\n"
        "['/IsCracked','S:No'],\n"
        "['/LawType','Enum_ToggleLawTypeGeoDependent:Fatigue Law'],\n"
        "['/DamageCalculationMethod','Enum_ToggleDamageCalculationMethod:LOCAL STRESS ANALYSIS'],\n"
        "['/FatigueLaw','Enum_ToggleFatigueLaw:AFI LAW'],\n"
        "['/Orientation_init','Enum_Orientation:" + campo20 + "'],\n"
        "['/Configuration_init','S:Configuration:AFI/thickness:50-200'],\n" #HAY QUE CAMBIAR ESTO
        "])\n\n")

        f.writelines("MS.CreateStandaloneAnalysis('initiation_lug',None,[\n"
        "   '/CsmMbr_Name S:" + keys + "',\n"
        "   '/Geometry *AirbusEO_DLugGeometry:',\n"
        "   '/Geometry/ReferencedObject AirbusEO_DLugGeometry:Geom" + str(cont) + "',\n"
        "   '/Loading *AirbusEO_DLugLoading:',\n"
        "   '/Loading/ReferencedObject AirbusEO_DLugLoading:Loading" + str(cont) + "',\n"
        "   '/FatigueLaw *AirbusEO_DFatigueLawGeoDependent:',\n"
        "   '/FatigueLaw/ReferencedObject AirbusEO_DFatigueLawGeoDependent:Law" + str(cont) + "',\n"
        "],\n"
        "'"+ keys + "')\n\n")

        f.writelines("\n\n")

        f.writelines("MS.RunAllAnalysis()\n")
        f.writelines("MS.Save('Lug.czm')\n")


        cont = cont + 1

