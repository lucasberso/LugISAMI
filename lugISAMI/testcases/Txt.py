import openpyxl
from path import Path
testcases_dir = Path(__file__).dirname()/"testcases"
if __name__ == '__main__':

    filename = '../H/Lug_manual.xlsm'
    book= openpyxl.load_workbook(testcases_dir+ '/'+ filename)
    sheet = book['Analysis']
    initial_row = 3
# final_row=sheet.max_row
    final_row = 4
    initial_column = 1
    final_column = sheet.max_column - 1

dict_analysis = {}

for row in range(initial_row, final_row):
    location = sheet.cell(row, 1).value
    dict = {}

    for column in range(initial_column + 1, final_column - 2):
        keyname = sheet.cell(2, column).value
        value = sheet.cell(row, column).value
        dict.update({keyname: value})

    dict_analysis.update({location: dict})



