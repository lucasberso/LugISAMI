#
# Authors:
# Pablo Arcediano,
# Lucas Bernacer Soriano,
# Ángela Blazquez,
# Héctor Dionisio,
# Javier Vela
#
# Copyright (c) 2021, Altran
#

import openpyxl
import os.path
from datetime import date

class Lug_generator():
    """
    Clase desarrolada para trabajar con orejetas en ISAMI.
    """

    # ----------------------- Inicialización
    def __init__(self, excel_filename, excel_path):
        """
        Inicializa la ruta, el nombre del fichero y lee el archivo Excel proporcionado.

        :str excel_filename: Nombre del fichero Excel de entrada.
        :str excel_path: Ruta del fichero Excel de entrada.
        """
        self.excel_filename = excel_filename
        self.excel_path = excel_path
        self.input_file = excel_path + '/' + excel_filename
        self.book = openpyxl.load_workbook(self.input_file, data_only=True)

    def read_input(self, initial_row=4, initial_column=2, header_row=3, name_sheet='Analysis'):
        """
        Obtiene de la hoja seleccionada por el usuario todos los datos encerrados por el rango suministrado.

        :int initial_row: Fila inicial de comienzo los datos.
        :int initial_column: Columna inicial de comienzo los datos.
        :int header_row: Fila con el nombre de los campos a almacenar.
        :int name_sheet: Nombre de la hoja de datos.
        :return: Diccionario de salida con los datos extraidos de la hoja Excel.
        """
        sheet = self.book[name_sheet]
        final_row = sheet.max_row
        final_column = sheet.max_column
        input_global = {}
        for i in range(initial_row, final_row + 1):
            name = sheet.cell(i, initial_column).value
            aux_dict = {}
            for j in range(initial_column, final_column + 1):
                key = sheet.cell(header_row, j).value
                value = sheet.cell(i, j).value
                aux_dict.update({key: value})
            input_global.update({name: aux_dict})
        return input_global

    def read_template(self):
        """
        Lee del archivo de entrada Excel los datos procedentes de las pestañas análisis y materiales.
        """
        self.analysis_data = self.read_input(initial_row = 4, initial_column = 1, header_row = 3, name_sheet='Analysis')
        self.material_data = self.read_input(initial_row = 5, initial_column = 2, header_row = 4, name_sheet='Materials')

    def write_output(self, output_filename):

        self.read_template()

        if os.path.isfile(output_filename + '.txt'):
            os.remove(output_filename + '.txt')
        file = open(output_filename + '.txt', "x")
        today_date = date.today()

        file.writelines("########################\n")
        file.writelines("# ISAMI VERSION: 8.0.0 #\n")
        file.writelines("# ANALYSIS: LUG        #\n")
        file.writelines("# Mode: SAA            #\n")
        file.writelines("# Written by: ALTRAN   #\n")
        file.writelines("# Date: " + today_date.strftime("%d/%m/%Y") + "     #\n")
        file.writelines("########################\n")

        for keys in self.material_data.keys():
            campo1 = keys
            campo2 = self.material_data[keys]["ISAMI Name"]
            campo3 = self.material_data[keys]["CODE"]
            campo4 = self.material_data[keys]["User/Referenced"]
            file.writelines("MS.LoadMaterial("+"'"+ campo1 +"'" +"," + campo2 +"," + campo3 +"," + campo4 +")\n")

        cont = 1
        for keys in self.analysis_data.keys():
            campo1 = self.analysis_data[keys]["Standar Pin"]
            campo2 = self.analysis_data[keys]["/Diameter"]
            campo3 = self.analysis_data[keys]["/Material"]
            campo4 = self.analysis_data[keys]["Standar Bush"]
            campo5 = self.analysis_data[keys]["/BushExternalDiameter"]
            campo6 = self.analysis_data[keys]["/BushMaterial"]
            campo7 = self.analysis_data[keys]["/LugResultantForce"]
            campo8 = self.analysis_data[keys]["/LugResultantAngle"]
            campo9 = self.analysis_data[keys]["/LoadingRatio"]
            campo10 = self.analysis_data[keys]["/Width"]
            campo11 = self.analysis_data[keys]["/Length"]
            campo12 = self.analysis_data[keys]["/Thick"]
            campo13 = self.analysis_data[keys]["/ThickUnstudied"]
            campo14 = self.analysis_data[keys]["/PinOffsetX"]
            campo15 = self.analysis_data[keys]["/PinOffsetY"]
            campo16 = self.analysis_data[keys]["/SuperiorAngle"]
            campo17 = self.analysis_data[keys]["/InferiorAngle"]
            campo18 = self.analysis_data[keys]["/ShearType"]
            campo19 = self.analysis_data[keys]["/StructureMaterial"]
            campo20 = self.analysis_data[keys]["/Orientation_init"]

            file.writelines("MS.CreateObject('DPines" + str(cont) + "','AirbusEO_DPin',[\n"
            "['/CsmMbr_Name','S:DPin" + str(cont) + "'],\n"
            "['/CsmMbr_Uptodate','B:TRUE'],\n"
            "['/Diameter','CaesamQty_LENGTH:" + str(campo2) + ";mm'],\n"
            "['/Material','AirbusEO_TMaterial:" + campo3 + "'],\n"
            "])\n\n")

            file.writelines("MS.CreateObject('DBushes" + str(cont) + "','AirbusEO_DBush',[\n"
            "['/CsmMbr_Name','S:D_Bush" + str(cont) + "'],\n"
            "['/BushInternalDiameter','CaesamQty_LENGTH:" + str(campo2) + ";mm'],\n"
            "['/BushExternalDiameter','CaesamQty_LENGTH:" + str(campo5) + ";mm'],\n"
            "['/BushMaterial','AirbusEO_TMaterial:" + campo6 + "'],\n"
            "])\n\n")

            file.writelines("MS.CreateObject('Geom" + str(cont) + "','AirbusEO_DLugGeometry',[\n"
            "['/CsmMbr_Name','S:Geom'],\n"
            "['/LugType','Enum_LugType:SIMPLE'],\n"
            "['/Width','CaesamQty_LENGTH:" + str(campo10) + ";mm'],\n"
            "['/Length','CaesamQty_LENGTH:" + str(campo11) + ";mm'],\n"
            "['/Thick','CaesamQty_LENGTH:" + str(campo12) + ";mm'],\n"
            "['/ThickUnstudied','CaesamQty_LENGTH:" + str(campo13) + ";mm'],\n"
            "['/LugPin','AirbusEO_DPin:DPines" + str(cont) + "'],\n"
            "['/LugBush','AirbusEO_DBush:DBushes" + str(cont) + "'],\n"
            "['/PinOffsetX','CaesamQty_LENGTH:" + str(campo14) + ";mm'],\n"
            "['/PinOffsetY','CaesamQty_LENGTH:" + str(campo15) + ";mm'],\n"
            "['/SuperiorAngle','CaesamQty_PLANE_ANGLE:" + str(campo16) + ";deg'],\n"
            "['/InferiorAngle','CaesamQty_PLANE_ANGLE:" + str(campo17) + ";deg'],\n"
            "['/NotchedLength','CaesamQty_LENGTH:20;mm'],\n"
            "['/ShearType','Enum_ToggleShearType:" + campo18 + "'],\n"
            "['/StructureMaterial','AirbusEO_TMaterial:" + campo19 + "'],\n"
            "])\n\n")

            file.writelines("MS.CreateObject('Loading" + str(cont) + "', 'AirbusEO_DLugLoading,[\n"
            "['/CsmMbr_Name','S:Loading'],\n"
            "['/LoadingType','Enum_LoadingType:NO_COEFFICIENT'],\n"
            "['/LugForceLoadingType','Enum_TogglePHForceLoadingType:RESULTANT AND ANGLE'],\n"
            "['/LugCoefficientForceX','D:20'],\n"
            "['/LugCoefficientForceY','D:30'],\n"
            "['/LugForceX','CaesamQty_FORCE:10000;N'],\n"
            "['/LugForceY','CaesamQty_FORCE:10000;N'],\n"
            "['/LugCoefficientResultantForce','D:100'],\n"
            "['/LugResultantForce','CaesamQty_FORCE:" + str(campo7) + ";N'],\n"
            "['/LugResultantAngle','CaesamQty_PLANE_ANGLE:" + str(campo8) + ";deg'],\n"
            "['/NbParameters','I:32'],\n"
            "['/LoadingRatio','D:" + str(campo9) + "'],\n"
            "['/NbClasses','I:1'],\n"
            "['/Compression','Enum_ToggleCompression:Smin'],\n"
            "['/UserSmin','CaesamQty_PRESSURE:0;MPa'],\n"
            "['/PropagationComputation','CaesamEnum_YesNo:No'],\n"
            "['/MLPDesign','CaesamEnum_YesNo:No'],\n"
            "['/LoadRedistributionFactor','D:1']\n"
            "]))\n\n")

            file.writelines("MS.CreateObject('Law" + str(cont) + "','AirbusEO_DFatigueLawGeoDependent',[\n"
            "['/CsmMbr_Name','S:Law'],\n"
            "['/IsCracked','S:No'],\n"
            "['/LawType','Enum_ToggleLawTypeGeoDependent:Fatigue Law'],\n"
            "['/DamageCalculationMethod','Enum_ToggleDamageCalculationMethod:LOCAL STRESS ANALYSIS'],\n"
            "['/FatigueLaw','Enum_ToggleFatigueLaw:AFI LAW'],\n"
            "['/Orientation_init','Enum_Orientation:" + campo20 + "'],\n"
                                                                                                                    "['/Configuration_init','S:Configuration:AFI/thickness:50-200'],\n"  # HAY QUE CAMBIAR ESTO
                                                                                                                    "])\n\n")

            file.writelines("MS.CreateStandaloneAnalysis('initiation_lug',None,[\n"
            "   '/CsmMbr_Name S:" + keys + "',\n"
            "   '/Geometry *AirbusEO_DLugGeometry:',\n"
            "   '/Geometry/ReferencedObject AirbusEO_DLugGeometry:Geom" + str(cont) + "',\n"
            "   '/Loading *AirbusEO_DLugLoading:',\n"
            "   '/Loading/ReferencedObject AirbusEO_DLugLoading:Loading" + str(cont) + "',\n"
            "   '/FatigueLaw *AirbusEO_DFatigueLawGeoDependent:',\n"
            "   '/FatigueLaw/ReferencedObject AirbusEO_DFatigueLawGeoDependent:Law" + str(cont) + "',\n"
            "],\n"
            "'" + keys + "')\n\n")

            file.writelines("\n\n")


            cont = cont + 1

        file.writelines("MS.RunAllAnalysis()\n")
        file.writelines("MS.Save('" + output_filename +".czm')\n")
        

    # def read_materials(self):
    #     """
    #     Función encargada de leer la pestaña de materiales en el libro Excel de entrada. Suministra un diccionario
    #     de salida que contiene todos los materiales y sus características.
    #
    #     """
    #     mat_sheet = self.book["Materials"]
    #     initial_row, final_row = 5, mat_sheet.max_row
    #     initial_column, final_column = 2, mat_sheet.max_column
    #     self.mat_global = {}
    #     for i in range(initial_row, final_row + 1):
    #         mat_name = mat_sheet.cell(i, initial_column).value
    #         aux_dict = {}
    #         for j in range(initial_column, final_column + 1):
    #             key = mat_sheet.cell(4, j).value
    #             value = mat_sheet.cell(i, j).value
    #             aux_dict.update({key: value})
    #         self.mat_global.update({mat_name: aux_dict})
    #     return self.mat_global
    # # def write_output(self):
    #
    #     # f = open("myfile.txt", "x")
    #     # f.writelines("########################\n")
    #     # f.writelines("# ISAMI VERSION: 8.0.0 #\n")
    #     # f.writelines("# ANALYSIS: LUG        #\n")
    #     # f.writelines("# Mode: SAA            #\n")
    #     # f.writelines("# Written by: ALTRAN   #\n")
    #     # f.writelines("# Date: 25/01/14       #\n")
    #     # f.writelines("########################\n")