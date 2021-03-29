#import os.path
#filename="ISAMI_FINAL.py"
#file_exists = os.path.isfile(filename)

#if file_exists:
#    pass
#else:
#    f = open("ISAMI_FINAL.py", "w")
#    f.write("Created file")

import openpyxl

path=
input = openpyxl.load_workbook('Lug_manual.xlsm', 'r')
ws = input.active
initial_row = 3
final_row = ws.max_row
initial_column = 2
final_column = ws.max_column - 1
dictio = {}
headers = ['Load', 'Angle']
if initial_row == final_row:
    vector = [initial_row]
else:
    vector = list(range(initial_row, final_row))
for i in vector:
    print(i)
    for j in range(initial_column, final_column):
        pass
print('Hola')



