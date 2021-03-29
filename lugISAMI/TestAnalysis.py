import openpyxl
from path import Path
testcases_dir = Path(__file__).dirname()/"testcases"
if __name__ == '__main__':

    filename = 'Lug_manual.xlsm'
    book= openpyxl.load_workbook(testcases_dir+ '/'+ filename, data_only = True)
    sheet = book["Analysis"]
    initial_row = 4
    final_row = sheet.max_row

    initial_column = 2
    final_column = sheet.max_column - 1
    cont = 1
    dict_global = {}
    for i in range(initial_row, final_row + 1):

        Location = sheet.cell(i, 1).value
        print(Location)

        dict = {}

        for j in range(initial_column, final_column + 1):
            keyname = sheet.cell(3, j).value
            value = sheet.cell(i, j).value
            dict.update({keyname: value})

        dict_global.update({Location: dict})

    f = open("myfile.txt", "x")
    f.writelines("########################\n")
    f.writelines("# ISAMI VERSION: 8.0.0 #\n")
    f.writelines("# ANALYSIS: LUG        #\n")
    f.writelines("# Mode: SAA            #\n")
    f.writelines("# Written by: ALTRAN   #\n")
    f.writelines("# Date: 25/01/14       #\n")
    f.writelines("########################\n")

    for keys in dict_global.keys():
        campo1 = dict_global[keys]["Standar Pin"]
        campo2 = dict_global[keys]["/Diameter"]
        campo3 = dict_global[keys]["/Material"]
        campo4 = dict_global[keys]["Standar Bush"]
        campo5 = dict_global[keys]["/BushExternalDiameter"]
        campo6 = dict_global[keys]["/BushMaterial"]
        campo7 = dict_global[keys]["/LugResultantForce"]
        campo8 = dict_global[keys]["/LugResultantAngle"]
        campo9 = dict_global[keys]["/LoadingRatio"]
        campo10 = dict_global[keys]["/Width"]
        campo11 = dict_global[keys]["/Length"]
        campo12 = dict_global[keys]["/Thick"]
        campo13 = dict_global[keys]["/ThickUnstudied"]
        campo14 = dict_global[keys]["/PinOffsetX"]
        campo15 = dict_global[keys]["/PinOffsetY"]
        campo16 = dict_global[keys]["/SuperiorAngle"]
        campo17 = dict_global[keys]["/InferiorAngle"]
        campo18 = dict_global[keys]["/ShearType"]
        campo19 = dict_global[keys]["/StructureMaterial"]
        campo20 = dict_global[keys]["/Orientation_init"]



        f.writelines("MS.CreateObject('DPines" + str(cont) + "','AirbusEO_DPin',[\n"
            "['/CsmMbr_Name','S:DPin" + str(cont) + "'],\n"
            "['/CsmMbr_Uptodate','B:TRUE'],\n"
            "['/Diameter','CaesamQty_LENGTH:" + str(campo2) + ";mm'],\n"
            "['/Material','AirbusEO_TMaterial:" + campo3 + "'],\n"
            "])\n\n")

        f.writelines("MS.CreateObject('DBushes" + str(cont) + "','AirbusEO_DBush',[\n"
            "['/CsmMbr_Name','S:D_Bush" + str(cont) + "'],\n"
            "['/BushInternalDiameter','CaesamQty_LENGTH:" + str(campo2) + ";mm'],\n"
            "['/BushExternalDiameter','CaesamQty_LENGTH:" + str(campo5) + ";mm'],\n"
            "['/BushMaterial','AirbusEO_TMaterial:" + campo6 + "'],\n"
            "])\n\n")


        cont = cont + 1

